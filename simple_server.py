#!/usr/bin/env python3
"""
Servidor Flask Simples - Gerador de IR
Versão limpa e funcional
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
import os
import re
import logging
from datetime import datetime
from pathlib import Path
# Import do gerador de PDF (opcional para funcionalidade básica)
try:
    import sys
    sys.path.append('Scripts')
    from gerador_ir_refatorado import GeradorIR
    PDF_GENERATOR_AVAILABLE = True
except ImportError as e:
    logger.warning(f"Gerador de PDF não disponível: {e}")
    PDF_GENERATOR_AVAILABLE = False

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Inicializar Flask
app = Flask(__name__)
CORS(app)

# Configurações
EXCEL_FILE = 'IR 2024 - NÃO ALTERAR.xlsx'
HOST = os.environ.get('HOST', '0.0.0.0')
PORT = int(os.environ.get('PORT', 10000))

class ExcelProcessor:
    """Processador simples do Excel"""
    
    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self._workbook = None
    
    def _load_workbook(self):
        """Carrega o workbook"""
        if self._workbook is None:
            if not self.file_path.exists():
                raise FileNotFoundError(f"Arquivo não encontrado: {self.file_path}")
            
            logger.info(f"Carregando Excel: {self.file_path}")
            self._workbook = load_workbook(self.file_path, data_only=True)
        
        return self._workbook
    
    def _normalize_cpf(self, cpf):
        """Normaliza CPF removendo caracteres especiais"""
        if not cpf:
            return ""
        # Remove tudo exceto números
        cpf_clean = re.sub(r'[^\d]', '', str(cpf))
        return cpf_clean
    
    def search_client(self, cpf):
        """Busca cliente por CPF"""
        try:
            wb = self._load_workbook()
            
            if 'Base de Clientes ' not in wb.sheetnames:
                logger.error("Planilha 'Base de Clientes ' não encontrada")
                return None
            
            ws = wb['Base de Clientes ']
            cpf_clean = self._normalize_cpf(cpf)
            
            logger.info(f"Buscando CPF normalizado: {cpf_clean}")
            
            # Buscar por CPF na coluna B (2) - onde estão os CPFs
            for row in range(2, ws.max_row + 1):
                cpf_cell = ws.cell(row=row, column=2).value  # Coluna B
                if cpf_cell:
                    cpf_cell_clean = self._normalize_cpf(cpf_cell)
                    
                    # Comparar CPFs normalizados
                    if cpf_clean == cpf_cell_clean:
                        nome = ws.cell(row=row, column=1).value  # Coluna A - Nome
                        empreendimento = ws.cell(row=row, column=3).value or 'N/A'  # Coluna C
                        
                        cliente = {
                            'cpf': cpf_clean,
                            'nome': nome,
                            'empreendimento': empreendimento
                        }
                        
                        logger.info(f"Cliente encontrado: {nome} (CPF: {cpf_clean})")
                        return cliente
            
            # Se não encontrou, mostrar alguns CPFs para debug
            logger.warning(f"CPF não encontrado: {cpf_clean}")
            logger.info("Primeiros CPFs na planilha:")
            for row in range(2, min(7, ws.max_row + 1)):
                cpf_cell = ws.cell(row=row, column=4).value  # Coluna D
                if cpf_cell:
                    cpf_cell_clean = self._normalize_cpf(cpf_cell)
                    logger.info(f"  Linha {row}: {cpf_cell} -> {cpf_cell_clean}")
            
            return None
            
        except Exception as e:
            logger.error(f"Erro ao buscar cliente: {str(e)}")
            return None
    
    def calculate_values(self, cpf):
        """Calcula valores financeiros"""
        try:
            wb = self._load_workbook()
            
            valores = {
                'receita_bruta': 0.0,
                'despesas_acessorias': 0.0,
                'saldo_union': 0.0,
                'saldo_paggo_dunning': 0.0
            }
            
            # Calcular valores da planilha UNION
            if 'UNION - 2024' in wb.sheetnames:
                ws_union = wb['UNION - 2024']
                valores['receita_bruta'] = self._sum_by_criteria(ws_union, cpf, "RECEITA BRUTA")
                valores['despesas_acessorias'] = self._sum_by_criteria(ws_union, cpf, "ATIVO CIRCULANTE")
                valores['saldo_union'] = valores['receita_bruta']
            
            # Calcular valores da planilha ERP
            if 'UNIFICADA ERP (paggo e dunning)' in wb.sheetnames:
                ws_erp = wb['UNIFICADA ERP (paggo e dunning)']
                valores['saldo_paggo_dunning'] = self._sum_erp_values(ws_erp, cpf)
            
            logger.info(f"Valores calculados para CPF {cpf}: {valores}")
            return valores
            
        except Exception as e:
            logger.error(f"Erro ao calcular valores: {str(e)}")
            return valores
    
    def _sum_by_criteria(self, ws, cpf, tipo):
        """Soma valores por critérios"""
        total = 0.0
        
        for row in range(2, ws.max_row + 1):
            cpf_col = ws.cell(row=row, column=5).value  # Coluna E
            tipo_col = ws.cell(row=row, column=16).value  # Coluna P
            valor_col = ws.cell(row=row, column=7).value  # Coluna G
            
            if (cpf_col and str(cpf) in str(cpf_col) and
                tipo_col and tipo in str(tipo_col) and
                valor_col and isinstance(valor_col, (int, float))):
                total += float(valor_col)
        
        return total
    
    def _sum_erp_values(self, ws, cpf):
        """Soma valores do ERP"""
        total = 0.0
        
        for row in range(2, ws.max_row + 1):
            cpf_col = ws.cell(row=row, column=6).value  # Coluna F
            valor_col = ws.cell(row=row, column=17).value  # Coluna Q
            
            if (cpf_col and str(cpf) in str(cpf_col) and
                valor_col and isinstance(valor_col, (int, float))):
                total += float(valor_col)
        
        return total

# Instância do processador
excel_processor = ExcelProcessor(EXCEL_FILE)

def validate_cpf(cpf):
    """Valida CPF"""
    if not cpf:
        return False, "CPF é obrigatório"
    
    cpf_clean = re.sub(r'[^\d]', '', str(cpf))
    
    if len(cpf_clean) != 11:
        return False, "CPF deve ter 11 dígitos"
    
    if len(set(cpf_clean)) == 1:
        return False, "CPF inválido"
    
    return True, cpf_clean

@app.route('/')
def index():
    """Página inicial"""
    return send_file('index.html')

@app.route('/styles.css')
def styles():
    """Arquivo CSS"""
    return send_file('styles.css')

@app.route('/script.js')
def script():
    """Arquivo JavaScript"""
    return send_file('script.js')

@app.route('/Imagens/<filename>')
def serve_images(filename):
    """Servir imagens"""
    return send_file(f'Imagens/{filename}')

@app.route('/api/health')
def health():
    """Health check"""
    try:
        wb = excel_processor._load_workbook()
        return jsonify({
            'success': True,
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'sheets': wb.sheetnames
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'status': 'unhealthy',
            'error': str(e)
        }), 500

@app.route('/api/buscar-e-gerar-pdf', methods=['POST'])
def buscar_e_gerar_pdf():
    """Busca cliente e calcula valores"""
    try:
        if not request.is_json:
            return jsonify({
                'success': False,
                'message': 'Content-Type deve ser application/json'
            }), 400
        
        data = request.get_json()
        cpf_raw = data.get('cpf', '')
        
        # Validar CPF
        is_valid, cpf_clean = validate_cpf(cpf_raw)
        if not is_valid:
            return jsonify({
                'success': False,
                'message': cpf_clean
            }), 400
        
        logger.info(f"Processando CPF: {cpf_clean}")
        
        # Buscar cliente
        cliente = excel_processor.search_client(cpf_clean)
        if not cliente:
            return jsonify({
                'success': False,
                'message': 'Cliente não encontrado na base de dados'
            }), 404
        
        # Calcular valores
        valores = excel_processor.calculate_values(cpf_clean)
        
        resultado = {
            'success': True,
            'message': 'Processamento concluído',
            'cliente': cliente,
            'valores': valores
        }
        
        logger.info(f"Processamento concluído para CPF: {cpf_clean}")
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro no processamento: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }), 500

@app.route('/api/gerar-pdf', methods=['POST'])
def gerar_pdf():
    """Gera PDF da declaração de IR"""
    try:
        if not request.is_json:
            return jsonify({
                'success': False,
                'message': 'Content-Type deve ser application/json'
            }), 400
        
        data = request.get_json()
        cpf_raw = data.get('cpf', '')
        
        # Validar CPF
        is_valid, cpf_clean = validate_cpf(cpf_raw)
        if not is_valid:
            return jsonify({
                'success': False,
                'message': cpf_clean
            }), 400
        
        logger.info(f"Gerando PDF para CPF: {cpf_clean}")
        
        # Gerar PDF
        if not PDF_GENERATOR_AVAILABLE:
            return jsonify({
                'success': False,
                'message': 'Gerador de PDF não disponível no momento'
            }), 503
        
        try:
            gerador = GeradorIR()
            sucesso, resultado = gerador.gerar_declaracao(cpf_clean)
            
            if sucesso:
                pdf_path = resultado
                if pdf_path and os.path.exists(pdf_path):
                    return send_file(
                        pdf_path,
                        as_attachment=True,
                        download_name=f"Declaracao_IR_{cpf_clean}.pdf",
                        mimetype='application/pdf'
                    )
                else:
                    return jsonify({
                        'success': False,
                        'message': 'PDF gerado mas arquivo não encontrado'
                    }), 500
            else:
                return jsonify({
                    'success': False,
                    'message': resultado
                }), 500
        except Exception as e:
            logger.error(f"Erro ao gerar PDF: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Erro ao gerar PDF: {str(e)}'
            }), 500
        
    except Exception as e:
        logger.error(f"Erro ao gerar PDF: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }), 500

@app.route('/api/test-simple', methods=['POST'])
def test_simple():
    """Teste simples"""
    try:
        data = request.get_json()
        cpf = data.get('cpf', '')
        
        return jsonify({
            'success': True,
            'message': 'Teste funcionando',
            'cpf_recebido': cpf,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro: {str(e)}'
        }), 500

if __name__ == '__main__':
    logger.info("🚀 Iniciando servidor simples...")
    logger.info(f"📁 Arquivo Excel: {EXCEL_FILE}")
    logger.info(f"🌐 Servidor: http://{HOST}:{PORT}")
    
    app.run(
        host=HOST,
        port=PORT,
        debug=False,
        threaded=True
    ) 