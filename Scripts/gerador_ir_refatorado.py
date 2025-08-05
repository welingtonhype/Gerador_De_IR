"""
Gerador de IR Refatorado - Versão Simplificada
Implementa exatamente as fórmulas Excel fornecidas
"""

import openpyxl
from openpyxl import load_workbook
import os
import re
import logging
from datetime import datetime
from pathlib import Path
import sys
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from config import get_config

config = get_config()

def setup_logging():
    """Configura logging"""
    log_dir = Path('logs')
    log_dir.mkdir(exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_dir / 'gerador_ir_refatorado.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

class ValidadorCPF:
    """Classe para validação de CPF"""
    
    @staticmethod
    def limpar_cpf(cpf):
        """Remove caracteres não numéricos do CPF"""
        if not cpf:
            return ""
        return re.sub(r'[^\d]', '', str(cpf))
    
    @staticmethod
    def validar_cpf(cpf):
        """Valida CPF com algoritmo oficial"""
        cpf_limpo = ValidadorCPF.limpar_cpf(cpf)
        
        if len(cpf_limpo) != 11:
            return False, "CPF deve ter 11 dígitos"
        
        if len(set(cpf_limpo)) == 1:
            return False, "CPF inválido"
        
        try:
            soma = sum(int(cpf_limpo[i]) * (10 - i) for i in range(9))
            resto = soma % 11
            digito1 = 0 if resto < 2 else 11 - resto
            
            soma = sum(int(cpf_limpo[i]) * (11 - i) for i in range(10))
            resto = soma % 11
            digito2 = 0 if resto < 2 else 11 - resto
            
            if int(cpf_limpo[9]) == digito1 and int(cpf_limpo[10]) == digito2:
                return True, cpf_limpo
            else:
                return False, "CPF inválido"
        except (ValueError, IndexError):
            return False, "CPF inválido"

class BuscadorCliente:
    """Classe para busca de dados do cliente - IMPLEMENTA FÓRMULAS EXCEL"""
    
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
    
    def buscar_por_cpf(self, cpf_busca):
        """
        Implementa: =PROCV(B27;'Base de Clientes'!C:D;2;FALSO)
        Busca CPF na Base de Clientes e retorna dados do cliente
        """
        is_valid, cpf_clean = ValidadorCPF.validar_cpf(cpf_busca)
        if not is_valid:
            logger.error(f"CPF inválido: {cpf_busca}")
            return None
        
        try:
            wb = load_workbook(self.arquivo_excel, data_only=True)
            
            if 'Base de Clientes ' not in wb.sheetnames:
                logger.error("Planilha 'Base de Clientes ' não encontrada")
                wb.close()
                return None
            
            ws = wb['Base de Clientes ']
            
            # Buscar CPF na coluna B (índice 2) - começando da linha 2 (dados reais)
            for row in range(2, ws.max_row + 1):
                cpf_cell = ws.cell(row=row, column=2).value  # Coluna B - CPF
                
                if cpf_cell:
                    cpf_cell_limpo = ValidadorCPF.limpar_cpf(cpf_cell)
                    if cpf_cell_limpo == cpf_clean:
                        # Extrair dados conforme estrutura da planilha (13 colunas)
                        dados = {
                            'cpf': cpf_clean,
                            'cliente': str(ws.cell(row=row, column=1).value or ''),  # Coluna A - Cliente
                            'empreendimento': str(ws.cell(row=row, column=3).value or ''),  # Coluna C - Empreendimento
                            'sigla': str(ws.cell(row=row, column=4).value or ''),  # Coluna D - Sigla
                            'unidade': str(ws.cell(row=row, column=5).value or ''),  # Coluna E - Unidade
                            'nome_social': str(ws.cell(row=row, column=6).value or ''),  # Coluna F - Nome Social
                            'cnpj_empreendimento': str(ws.cell(row=row, column=7).value or ''),  # Coluna G - CNPJ Empreendimento
                            'endereco': str(ws.cell(row=row, column=8).value or ''),  # Coluna H - Endereço
                            'numero': str(ws.cell(row=row, column=9).value or ''),  # Coluna I - Número
                            'bairro': str(ws.cell(row=row, column=10).value or ''),  # Coluna J - Bairro
                            'cidade': str(ws.cell(row=row, column=12).value or ''),  # Coluna L - Cidade
                            'estado': str(ws.cell(row=row, column=11).value or ''),  # Coluna K - Estado
                            'valor_venda': self._converter_valor_venda(ws.cell(row=row, column=13).value)  # Coluna M - Valor de Venda
                        }
                        
                        wb.close()
                        logger.info(f"Cliente encontrado: {dados['cliente']} - Empreendimento: {dados['empreendimento']}")
                        return dados
            
            wb.close()
            logger.warning(f"CPF não encontrado: {cpf_clean}")
            return None
            
        except Exception as e:
            logger.error(f"Erro ao buscar cliente: {str(e)}")
            return None
    
    def _converter_valor_venda(self, valor):
        """Converte o valor da venda para float, tratando casos especiais"""
        if valor is None:
            return 0
        
        valor_str = str(valor).strip()
        
        # Se for "Verificar" ou similar, retorna 0
        if valor_str.lower() in ['verificar', 'verificar ', 'n/a', '', 'verificar']:
            return 0
        
        try:
            return float(valor_str)
        except (ValueError, TypeError):
            return 0

class CalculadorFinanceiro:
    """Classe para cálculos financeiros - IMPLEMENTA FÓRMULAS EXCEL"""
    
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
    
    def calcular_receita_bruta(self, cpf_cliente):
        """
        Implementa: =SOMASES('UNION - 2024'!G:G;'UNION - 2024'!P:P;"RECEITA BRUTA";'UNION - 2024'!E:E;Declaração!B27)
        G:G = ENTRADA (coluna 3)
        P:P = DIVISÃO - 1º NÍVEL (coluna 4) 
        E:E = CLIENTE (coluna 2)
        """
        try:
            wb = load_workbook(self.arquivo_excel, data_only=True)
            
            if 'UNION - 2024' not in wb.sheetnames:
                logger.error("Planilha 'UNION - 2024' não encontrada")
                wb.close()
                return 0
            
            # Primeiro, buscar o nome do cliente na Base de Clientes
            if 'Base de Clientes ' not in wb.sheetnames:
                logger.error("Planilha 'Base de Clientes ' não encontrada")
                wb.close()
                return 0
            
            ws_base = wb['Base de Clientes ']
            nome_cliente = None
            
            # Buscar nome do cliente por CPF
            for row in range(2, ws_base.max_row + 1):
                cpf_cell = ws_base.cell(row=row, column=2).value  # Coluna B - CPF
                
                if cpf_cell:
                    cpf_cell_limpo = re.sub(r'[^\d]', '', str(cpf_cell))
                    cpf_cliente_limpo = re.sub(r'[^\d]', '', str(cpf_cliente))
                    
                    if cpf_cell_limpo == cpf_cliente_limpo:
                        nome_cliente = ws_base.cell(row=row, column=1).value  # Coluna A - Cliente
                        break
            
            if not nome_cliente:
                logger.warning(f"Nome do cliente não encontrado para CPF: {cpf_cliente}")
                wb.close()
                return 0
            
            logger.info(f"Buscando dados para cliente: {nome_cliente}")
            
            # Agora buscar na UNION - 2024 pelo nome do cliente
            ws = wb['UNION - 2024']
            total = 0
            
            # Buscar por nome do cliente na coluna B e tipo "RECEITA BRUTA" na coluna D
            for row in range(2, ws.max_row + 1):
                cliente_col_b = ws.cell(row=row, column=2).value  # Coluna B - CLIENTE
                tipo_col_d = ws.cell(row=row, column=4).value  # Coluna D - DIVISÃO - 1º NÍVEL
                valor_col_c = ws.cell(row=row, column=3).value  # Coluna C - ENTRADA
                
                if (cliente_col_b and nome_cliente.lower() in str(cliente_col_b).lower() and
                    tipo_col_d and "RECEITA BRUTA" in str(tipo_col_d).upper() and
                    valor_col_c and isinstance(valor_col_c, (int, float))):
                    total += float(valor_col_c)
                    logger.debug(f"Receita bruta encontrada: R$ {valor_col_c:,.2f} - Cliente: {cliente_col_b}")
            
            wb.close()
            logger.info(f"Receita bruta total: R$ {total:,.2f}")
            return total
            
        except Exception as e:
            logger.error(f"Erro ao calcular receita bruta: {str(e)}")
            return 0
    
    def calcular_despesas_acessorias(self, cpf_cliente):
        """
        Implementa: =SOMASES('UNION - 2024'!G:G,'UNION - 2024'!P:P,"ATIVO CIRCULANTE",'UNION - 2024'!E:E,Declaração!B27)
        G:G = ENTRADA (coluna 3)
        P:P = DIVISÃO - 1º NÍVEL (coluna 4)
        E:E = CLIENTE (coluna 2)
        """
        try:
            wb = load_workbook(self.arquivo_excel, data_only=True)
            
            if 'UNION - 2024' not in wb.sheetnames:
                logger.error("Planilha 'UNION - 2024' não encontrada")
                wb.close()
                return 0
            
            # Primeiro, buscar o nome do cliente na Base de Clientes
            if 'Base de Clientes ' not in wb.sheetnames:
                logger.error("Planilha 'Base de Clientes ' não encontrada")
                wb.close()
                return 0
            
            ws_base = wb['Base de Clientes ']
            nome_cliente = None
            
            # Buscar nome do cliente por CPF
            for row in range(2, ws_base.max_row + 1):
                cpf_cell = ws_base.cell(row=row, column=2).value  # Coluna C - CPF
                
                if cpf_cell:
                    cpf_cell_limpo = re.sub(r'[^\d]', '', str(cpf_cell))
                    cpf_cliente_limpo = re.sub(r'[^\d]', '', str(cpf_cliente))
                    
                    if cpf_cell_limpo == cpf_cliente_limpo:
                        nome_cliente = ws_base.cell(row=row, column=1).value  # Coluna B - Cliente
                        break
            
            if not nome_cliente:
                logger.warning(f"Nome do cliente não encontrado para CPF: {cpf_cliente}")
                wb.close()
                return 0
            
            # Agora buscar na UNION - 2024 pelo nome do cliente
            ws = wb['UNION - 2024']
            total = 0
            
            # Buscar por nome do cliente na coluna B e tipo "ATIVO CIRCULANTE" na coluna D
            for row in range(2, ws.max_row + 1):
                cliente_col_b = ws.cell(row=row, column=2).value  # Coluna B - CLIENTE
                tipo_col_d = ws.cell(row=row, column=4).value  # Coluna D - DIVISÃO - 1º NÍVEL
                valor_col_c = ws.cell(row=row, column=3).value  # Coluna C - ENTRADA
                
                if (cliente_col_b and nome_cliente.lower() in str(cliente_col_b).lower() and
                    tipo_col_d and "ATIVO CIRCULANTE" in str(tipo_col_d).upper() and
                    valor_col_c and isinstance(valor_col_c, (int, float))):
                    total += float(valor_col_c)
                    logger.debug(f"Despesa acessória encontrada: R$ {valor_col_c:,.2f} - Cliente: {cliente_col_b}")
            
            wb.close()
            logger.info(f"Despesas acessórias total: R$ {total:,.2f}")
            return total
            
        except Exception as e:
            logger.error(f"Erro ao calcular despesas acessórias: {str(e)}")
            return 0
    
    def calcular_saldo_union(self, cpf_cliente):
        """
        Implementa: =SOMASES('UNION - 2024'!G:G;'UNION - 2024'!P:P;"RECEITA BRUTA";'UNION - 2024'!E:E;C3)
        Calcula saldo Union baseado no CPF
        """
        return self.calcular_receita_bruta(cpf_cliente)  # Mesma lógica da receita bruta

class GeradorPDF:
    """Classe para geração de PDF - MANTIDA COMO ESTAVA"""
    
    def __init__(self):
        self.config = config
    
    def _criar_estilos(self):
        """Cria estilos para o PDF"""
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=14,
            spaceAfter=5,
            alignment=TA_CENTER,
            textColor=colors.black,
            leading=16
        )
        
        section_style = ParagraphStyle(
            'SectionStyle',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=10,
            spaceAfter=8,
            spaceBefore=15,
            textColor=colors.black,
            leading=12
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            spaceAfter=6,
            textColor=colors.black,
            leading=12
        )
        
        return title_style, section_style, normal_style
    
    def _criar_cabecalho(self):
        """Cria cabeçalho com logos da Hype e Ministério da Fazenda"""
        try:
            # Verificar se as imagens existem
            logo_hype_path = "Imagens/Imagem2.png"
            logo_ministerio_path = "Imagens/Imagem1.png"
            
            if not os.path.exists(logo_hype_path):
                logger.warning(f"Logo da Hype não encontrado: {logo_hype_path}")
                logo_hype = None
            else:
                logo_hype = Image(logo_hype_path, width=1.2*inch, height=0.8*inch)
            
            if not os.path.exists(logo_ministerio_path):
                logger.warning(f"Logo do Ministério não encontrado: {logo_ministerio_path}")
                logo_ministerio = None
            else:
                logo_ministerio = Image(logo_ministerio_path, width=0.5*inch, height=0.5*inch)
            
            # Criar texto central
            texto_central = Paragraph(
                """<para align=center>
                <b>ANO-CALENDÁRIO DE 2024<br/>
                IMPOSTO DE RENDA - PESSOA FÍSICA</b>
                </para>""",
                ParagraphStyle(
                    'HeaderCenter',
                    fontName='Helvetica-Bold',
                    fontSize=11,
                    alignment=TA_CENTER,
                    textColor=colors.black,
                    leading=13
                )
            )
            
            # Criar seção direita com texto e brasão lado a lado
            if logo_ministerio:
                texto_ministerio = Paragraph(
                    """<para align=center>
                    <b>MINISTÉRIO DA<br/>
                    FAZENDA<br/>
                    SECRETARIA<br/>
                    DA<br/>
                    RECEITA FEDERAL</b>
                    </para>""",
                    ParagraphStyle(
                        'HeaderRight',
                        fontName='Helvetica-Bold',
                        fontSize=8,
                        alignment=TA_CENTER,
                        textColor=colors.black,
                        leading=10
                    )
                )
                
                # Tabela interna para texto + brasão
                ministerio_data = [[texto_ministerio, logo_ministerio]]
                ministerio_table = Table(ministerio_data, colWidths=[1.3*inch, 0.6*inch])
                ministerio_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ]))
                
                if logo_hype:
                    header_data = [[logo_hype, texto_central, ministerio_table]]
                    col_widths = [1.2*inch, 3.9*inch, 1.9*inch]
                else:
                    header_data = [[texto_central, ministerio_table]]
                    col_widths = [5.1*inch, 1.9*inch]
            else:
                texto_direito = Paragraph(
                    """<para align=center>
                    <b>MINISTÉRIO DA<br/>
                    FAZENDA<br/>
                    SECRETARIA<br/>
                    DA<br/>
                    RECEITA FEDERAL</b>
                    </para>""",
                    ParagraphStyle(
                        'HeaderRight',
                        fontName='Helvetica-Bold',
                        fontSize=8,
                        alignment=TA_CENTER,
                        textColor=colors.black,
                        leading=10
                    )
                )
                
                if logo_hype:
                    header_data = [[logo_hype, texto_central, texto_direito]]
                    col_widths = [1.2*inch, 4.5*inch, 1.3*inch]
                else:
                    header_data = [[texto_central, texto_direito]]
                    col_widths = [5.7*inch, 1.3*inch]
            
            # Criar tabela principal do cabeçalho
            header_table = Table(header_data, colWidths=col_widths)
            header_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            return header_table
            
        except Exception as e:
            logger.error(f"Erro ao criar cabeçalho: {str(e)}")
            return Paragraph(
                "<para align=center><b>ANO-CALENDÁRIO DE 2024<br/>IMPOSTO DE RENDA - PESSOA FÍSICA</b></para>",
                ParagraphStyle(
                    'SimpleHeader',
                    fontName='Helvetica-Bold',
                    fontSize=14,
                    alignment=TA_CENTER,
                    textColor=colors.black,
                    leading=16
                )
            )
    
    def _criar_tabela_pj(self, dados_cliente):
        """Cria tabela da pessoa jurídica"""
        nome_empresa = "HYPE EMPREENDIMENTOS"
        cnpj_empresa = "41.081.989/0001-92"
        
        pj_data = [[f"Nome Empresarial: {nome_empresa} - {cnpj_empresa}"]]
        pj_table = Table(pj_data, colWidths=[7*inch])
        pj_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, 0), 0.5, colors.black),
            ('MINIMUMHEIGHT', (0, 0), (-1, 0), 25),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, 0), 8),
            ('RIGHTPADDING', (0, 0), (-1, 0), 8),
        ]))
        return pj_table
    
    def _criar_tabela_fonte(self, dados_cliente):
        """Cria tabela da fonte pagadora"""
        cliente_nome = dados_cliente.get('cliente', '')
        cliente_cpf = dados_cliente.get('cpf', '')
        
        fonte_data = [[cliente_nome, 'CPF:', cliente_cpf]]
        fonte_table = Table(fonte_data, colWidths=[5.5*inch, 0.5*inch, 1*inch])
        fonte_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (2, 0), (2, 0), 'LEFT'),
            ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.black),
            ('LINEBEFORE', (0, 0), (0, 0), 0.5, colors.black),
            ('LINEAFTER', (-1, 0), (-1, 0), 0.5, colors.black),
            ('MINIMUMHEIGHT', (0, 0), (-1, 0), 25),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, 0), 8),
            ('RIGHTPADDING', (0, 0), (-1, 0), 8),
        ]))
        return fonte_table
    
    def _criar_tabela_bem(self, dados_cliente):
        """Cria tabela dos dados do bem"""
        endereco = dados_cliente.get('endereco', '')
        numero = dados_cliente.get('numero', '')
        bairro = dados_cliente.get('bairro', '')
        cidade = dados_cliente.get('cidade', '')
        estado = dados_cliente.get('estado', '')
        endereco_completo = f"{endereco}, {numero} - {bairro}, {cidade} - {estado}"
        
        sigla = dados_cliente.get('sigla', '')
        unidade = dados_cliente.get('unidade', '')
        
        # Montar produto com sigla + unidade
        if sigla and unidade:
            produto_texto = f"{sigla} - {unidade}"
        elif sigla:
            produto_texto = sigla
        elif unidade:
            produto_texto = unidade
        else:
            produto_texto = "Verificar"
        
        # Valor do imóvel
        valor_venda = dados_cliente.get('valor_venda', 0)
        if valor_venda and isinstance(valor_venda, (int, float)) and valor_venda > 0:
            valor_imovel = f"R$ {valor_venda:,.2f}"
        else:
            valor_imovel = "Verificar"
        
        bem_data = [
            ['Produto', produto_texto],
            ['Endereço', endereco_completo],
            ['Valor do Imóvel', valor_imovel]
        ]
        bem_table = Table(bem_data, colWidths=[2*inch, 5*inch])
        bem_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 0.5, colors.black),
            ('LINEBEFORE', (0, 0), (0, -1), 0.5, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 0.5, colors.black),
            ('MINIMUMHEIGHT', (0, 0), (-1, -1), 25),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        return bem_table
    
    def _criar_tabela_pagamentos(self, valores_calculados):
        """Cria tabela de pagamentos"""
        receita_bruta = valores_calculados.get('receita_bruta', 0)
        despesas_acessorias = valores_calculados.get('despesas_acessorias', 0)
        
        pagamentos_data = [
            ['ESPECIFICAÇÃO', 'VALORES PAGOS EM 2024'],
            ['RECEITA', f"R$ {receita_bruta:,.2f}"],
            ['DESPESAS ACESSÓRIAS', f"R$ {despesas_acessorias:,.2f}"]
        ]
        
        pagamentos_table = Table(pagamentos_data, colWidths=[4*inch, 3*inch])
        pagamentos_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        return pagamentos_table
    
    def gerar_declaracao(self, cpf, dados_cliente, valores_calculados):
        """Gera PDF da declaração de IR"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nome_pdf = f"Declaracao_IR_{cpf}_{timestamp}.pdf"
            
            doc = SimpleDocTemplate(nome_pdf, pagesize=A4, 
                                  leftMargin=0.3*inch, rightMargin=0.8*inch,
                                  topMargin=0.8*inch, bottomMargin=0.8*inch)
            story = []
            
            title_style, section_style, normal_style = self._criar_estilos()
            
            # Cabeçalho com logos
            header_table = self._criar_cabecalho()
            story.append(header_table)
            story.append(Spacer(1, 20))
            
            # Seção 1 - PESSOA JURÍDICA
            story.append(Paragraph("1. PESSOA JURÍDICA:", section_style))
            story.append(self._criar_tabela_pj(dados_cliente))
            story.append(Spacer(1, 15))
            
            # Seção 2 - FONTE PAGADORA PESSOA FÍSICA
            story.append(Paragraph("2. FONTE PAGADORA PESSOA FÍSICA:", section_style))
            story.append(self._criar_tabela_fonte(dados_cliente))
            story.append(Spacer(1, 15))
            
            # Seção 3 - DADOS DO BEM
            story.append(Paragraph("3. DADOS DO BEM:", section_style))
            story.append(self._criar_tabela_bem(dados_cliente))
            story.append(Spacer(1, 15))
            
            # Seção 4 - INFORME DE PAGAMENTOS
            story.append(Paragraph("4. INFORME DE PAGAMENTOS EFETUADOS PARA FINS DE IMPOSTO DE RENDA:", section_style))
            story.append(self._criar_tabela_pagamentos(valores_calculados))
            story.append(Spacer(1, 30))
            
            # Footer
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#D9D9D9'), spaceAfter=10))
            
            data_emissao = datetime.now().strftime('%d de %B de %Y')
            meses_pt = {
                'January': 'janeiro', 'February': 'fevereiro', 'March': 'março',
                'April': 'abril', 'May': 'maio', 'June': 'junho',
                'July': 'julho', 'August': 'agosto', 'September': 'setembro',
                'October': 'outubro', 'November': 'novembro', 'December': 'dezembro'
            }
            
            for mes_en, mes_pt in meses_pt.items():
                data_emissao = data_emissao.replace(mes_en, mes_pt)
            
            story.append(Paragraph(f"Emitido em {data_emissao}", normal_style))
            story.append(Spacer(1, 15))
            
            # Linha de assinatura centralizada
            story.append(HRFlowable(width="50%", thickness=0.5, color=colors.black, spaceAfter=5))
            
            # Texto da empresa centralizado embaixo da linha
            empresa_style = ParagraphStyle(
                'EmpresaStyle',
                parent=normal_style,
                fontName='Helvetica',
                fontSize=10,
                alignment=TA_CENTER,
                textColor=colors.black,
                leading=12
            )
            story.append(Paragraph("Hyperion Empreendimentos e Incorporações SA", empresa_style))
            
            doc.build(story)
            
            logger.info(f"PDF gerado com sucesso: {nome_pdf}")
            return nome_pdf
            
        except Exception as e:
            import traceback
            logger.error(f"Erro ao gerar PDF: {str(e)}")
            logger.error(f"Traceback completo: {traceback.format_exc()}")
            return None

class GeradorIR:
    """Classe principal do gerador de IR - VERSÃO SIMPLIFICADA"""
    
    def __init__(self):
        from config import FILES_CONFIG
        self.config = FILES_CONFIG
        self.arquivo_excel = FILES_CONFIG['EXCEL_FILE']
        self.buscador = BuscadorCliente(self.arquivo_excel)
        self.calculador = CalculadorFinanceiro(self.arquivo_excel)
        self.gerador_pdf = GeradorPDF()
    
    def gerar_declaracao(self, cpf):
        """Função principal para gerar declaração de IR"""
        logger.info(f"Iniciando geração de declaração para CPF: {cpf}")
        
        # Validar CPF
        is_valid, cpf_clean = ValidadorCPF.validar_cpf(cpf)
        if not is_valid:
            logger.error(f"CPF inválido: {cpf}")
            return False, "CPF inválido"
        
        # Buscar dados do cliente
        dados_cliente = self.buscador.buscar_por_cpf(cpf_clean)
        if not dados_cliente:
            logger.error(f"Cliente não encontrado para CPF: {cpf_clean}")
            return False, "Cliente não encontrado"
        
        logger.info(f"Cliente encontrado: {dados_cliente['cliente']}")
        
        # Calcular valores financeiros
        receita_bruta = self.calculador.calcular_receita_bruta(cpf_clean)
        despesas_acessorias = self.calculador.calcular_despesas_acessorias(cpf_clean)
        
        valores_calculados = {
            'receita_bruta': receita_bruta,
            'despesas_acessorias': despesas_acessorias
        }
        
        logger.info(f"Valores calculados - Receita: R$ {receita_bruta:,.2f}, Despesas: R$ {despesas_acessorias:,.2f}")
        
        # Gerar PDF
        nome_pdf = self.gerador_pdf.gerar_declaracao(cpf_clean, dados_cliente, valores_calculados)
        
        if nome_pdf:
            logger.info(f"Declaração gerada com sucesso: {nome_pdf}")
            return True, nome_pdf
        else:
            logger.error("Erro ao gerar PDF")
            return False, "Erro ao gerar PDF"

def main():
    """Função principal"""
    print("🏢 GERADOR DE DECLARAÇÃO DE IR - VERSÃO SIMPLIFICADA")
    print("=" * 60)
    
    gerador = GeradorIR()
    
    while True:
        print(f"\n📋 Opções:")
        print(f"   1. Gerar declaração para CPF específico")
        print(f"   2. Testar com CPF da planilha")
        print(f"   3. Sair")
        
        try:
            opcao = input(f"\nEscolha uma opção (1-3): ").strip()
            
            if opcao == "1":
                cpf = input(f"Digite o CPF: ").strip()
                if cpf:
                    sucesso, resultado = gerador.gerar_declaracao(cpf)
                    if sucesso:
                        print(f"✅ Declaração gerada com sucesso: {resultado}")
                    else:
                        print(f"❌ Erro: {resultado}")
            
            elif opcao == "2":
                from config import FILES_CONFIG
                cpf_teste = FILES_CONFIG['TEST']['TEST_CPF']
                print(f"Testando com CPF: {cpf_teste}")
                sucesso, resultado = gerador.gerar_declaracao(cpf_teste)
                if sucesso:
                    print(f"✅ Declaração gerada com sucesso: {resultado}")
                else:
                    print(f"❌ Erro: {resultado}")
            
            elif opcao == "3":
                print(f"👋 Saindo...")
                break
            
            else:
                print(f"❌ Opção inválida!")
                
        except KeyboardInterrupt:
            print(f"\n👋 Saindo...")
            break
        except Exception as e:
            logger.error(f"Erro no menu principal: {str(e)}")
            print(f"❌ Erro: {str(e)}")

if __name__ == "__main__":
    main() 